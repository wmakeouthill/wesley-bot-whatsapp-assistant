import logging
import uuid
import base64
import io
import asyncio
import re
import unicodedata
from typing import Optional
from datetime import datetime

import openpyxl
from gtts import gTTS
from google import genai
from sqlalchemy import select, or_

from app.domain.schemas.webhook import WebhookBody
from app.infrastructure.external.evolution_client import EvolutionClient
from app.domain.services.rag_service import PortfolioRAG
from app.domain.services.document_catalog_service import DocumentCatalogService, DocumentEntry
from app.infrastructure.database.session import async_session
from app.domain.entities.models import Cliente, Mensagem, BotConfig, AllowBlockEntry
from app.infrastructure.config.settings import settings

logger = logging.getLogger(__name__)

ASSISTANT_DISPLAY_NAME = "a assistente virtual do Wesley"
WESLEY_PUBLIC_NAME = "Wesley"
WESLEY_PUBLIC_FULL_NAME = "Wesley de Carvalho Augusto Correia"


# ---------------------------------------------------------------------------
# Prompts por "personalidade" do bot
# ---------------------------------------------------------------------------

PROMPT_PORTFOLIO = """
Você é o Assistente Virtual Oficial do Wesley no WhatsApp.
Seja amigável, direto, e natural em suas respostas.
Nome de exibição no WhatsApp: {nome_cliente}. Esse nome pode estar desatualizado ou incompleto.

Abaixo está o CONTEXTO contendo as informações sobre os certificados, habilidades e currículo do Wesley.
Baseado ESTREITAMENTE nesse contexto, responda a pergunta do usuário.
Se a informação não estiver no contexto, diga que você vai anotar a dúvida para o próprio Wesley responder depois, e NUNCA invente informações.

Regras críticas:
- Se já houver saudação anterior no histórico, não cumprimente de novo e não reabra a conversa com "olá", "oi", "tudo bem" ou equivalente.

CONTEXTO DEDUZIDO DO PORTFÓLIO:
{contexto}

{historico_prompt}ÚLTIMA MENSAGEM DO USUÁRIO: "{texto}"
"""

PROMPT_PORTFOLIO_AUDIO = """
Você é o Assistente Virtual Oficial do Wesley no WhatsApp.
Nome de exibição no WhatsApp: {nome_cliente}. Esse nome pode estar desatualizado ou incompleto.
Sua resposta será convertida em voz (TTS), portanto:
1. Seja extremamente natural, coloquial e amigável.
2. EVITE QUALQUER formatação markdown (sem asteriscos **, sem listas com -, sem hashtags #).
3. Escreva números e siglas de forma pronunciável (ex: "C Sharp" em vez de "C#", "Node J S" em vez de "Node.js", "cem por cento").
4. Seja mais objetivo e direto, fale como um áudio de WhatsApp real.

Abaixo está o CONTEXTO contendo as informações sobre o Wesley.
Baseado ESTREITAMENTE nesse contexto, responda a pergunta do usuário. Se não souber dizer, fale que não sabe. NUNCA invente informações.

Regras críticas:
- Se já houver saudação anterior no histórico, não cumprimente de novo e não reabra a conversa com "olá", "oi", "tudo bem" ou equivalente.

CONTEXTO DEDUZIDO DO PORTFÓLIO:
{contexto}

{historico_prompt}ÚLTIMA MENSAGEM DO USUÁRIO: "{texto}"
"""

PROMPT_PESSOAL = """
Você é a Assistente Virtual do WhatsApp pessoal do Wesley.
Você responde em nome dele, de forma cortês, natural e prestativa.

Regras de personalidade:
- Seja educada e cortês, mantendo um tom natural de WhatsApp. Seja menos informal, sem gírias exageradas (evite excesso de "kkk" ou risadas falsas).
- Se houver histórico de conversa do próprio Wesley, tente imitar sutilmente o estilo sem soar forçado, mas lembre-se que você é a assistente dele.
- Seja direta e objetiva, como num chat real.
- Não use markdown (sem **, sem #, sem listas com -).
- Se não souber algo, diga que vai anotar para o Wesley responder pessoalmente em breve. NUNCA invente informações.
- Nome de exibição no WhatsApp: {nome_cliente}. Esse nome pode estar desatualizado ou incompleto.
- Se o histórico já mostrar uma saudação anterior da assistente, não cumprimente de novo e não reabra a conversa com "olá", "oi", "tudo bem" ou equivalente.

Abaixo está o CONTEXTO contendo as informações profissionais, habilidades e portfólio do Wesley.
Sempre que perguntado sobre trabalho ou habilidades dele, responda baseando-se EXCLUSIVAMENTE nesse contexto:
{contexto}

HISTÓRICO RECENTE DA CONVERSA (para referência):
{historico}

MENSAGEM RECEBIDA DE {nome_cliente}: "{texto}"

Escreva apenas a resposta direta, de forma natural.
"""

PROMPT_PESSOAL_AUDIO = """
Você é a Assistente Virtual do WhatsApp pessoal do Wesley. Sua resposta será enviada como ÁUDIO.
Você responde em nome dele de forma cortês, educada e natural de WhatsApp.

Regras:
1. Seja cortês e amigável, sem informalidade exagerada.
2. EVITE QUALQUER formatação markdown.
3. Escreva para ser falado (ex: "Node J S", "C Sharp", "cem por cento").
4. Não invente informações. Se não souber, diga que o Wesley responde depois.
5. Nome de exibição no WhatsApp: {nome_cliente}. Esse nome pode estar desatualizado ou incompleto.
6. Se o histórico já mostrar uma saudação anterior da assistente, não cumprimente de novo e não reabra a conversa com "olá", "oi", "tudo bem" ou equivalente.

Abaixo está o CONTEXTO profissional do Wesley. Use essas informações se perguntarem de trabalho ou habilidades dele. NUNCA invente:
{contexto}

HISTÓRICO DA CONVERSA:
{historico}

MENSAGEM DE {nome_cliente}: "{texto}"
"""

# ---------------------------------------------------------------------------
# Comandos /ia disponíveis (ajuda inline)
# ---------------------------------------------------------------------------
AJUDA_COMANDOS = """🤖 *Comandos disponíveis:*

`/ia on` — ativa a IA para todos
`/ia off` — desativa a IA para todos
`/ia on 5521999999999` — ativa para número específico
`/ia off 5521999999999` — desativa para número específico
`/ia lista` — lista últimas 10 conversas com status
`/ia resetar 5521999999999` — remove override individual (volta ao padrão global)
`/ia status` — mostra se a IA está ativa globalmente nesta instância
"""


class AtendimentoService:
    """Serviço Orquestrador do fluxo conversacional"""

    def __init__(self, evolution_client: EvolutionClient):
        self.evolution_client = evolution_client
        self.rag = PortfolioRAG()
        self._rag_ready = False
        self._rag_init_lock = asyncio.Lock()
        self.document_catalog = DocumentCatalogService()
        self.llm_client = genai.Client(api_key=settings.gemini_api_key)

    async def ensure_rag_ready(self) -> None:
        if self._rag_ready:
            return
        async with self._rag_init_lock:
            if self._rag_ready:
                return
            logger.info("Inicializando RAG em background/lazy...")
            await asyncio.to_thread(self.rag.initialize_or_build)
            self._rag_ready = True
            logger.info("RAG pronto.")

    # -----------------------------------------------------------------------
    # Ponto de entrada do Webhook
    # -----------------------------------------------------------------------

    async def processar_webhook(self, body: WebhookBody, client: Optional[EvolutionClient] = None) -> None:
        """Ponto de entrada do Webhook. `client` permite override do EvolutionClient por instância."""
        ev_client = client or self.evolution_client
        instancia = body.instance

        if body.event != "messages.upsert":
            return

        remote_jid = body.data.key.remoteJid

        # --- Determina o owner_jid desta instância ---
        if instancia == settings.evolution_instance_two_name:
            owner_jid = settings.instance_two_owner_jid
        else:
            owner_jid = settings.owner_jid

        # --- Mensagens fromMe: só processar se for comando do owner ---
        if body.data.key.fromMe:
            # Verifica se a mensagem foi enviada para o próprio número (Mensagens Salvas)
            # ou se o owner está mandando mensagem - este é o canal de comandos /ia
            texto_cmd = self._extrair_texto(body)
            if texto_cmd and texto_cmd.strip().startswith("/ia"):
                await self._processar_comando_ia(texto_cmd.strip(), instancia, owner_jid, ev_client)
            return

        # --- Ignora grupos ---
        if "@g.us" in remote_jid:
            logger.info(f"Mensagem de grupo ignorada: {remote_jid}")
            return

        # --- Normaliza o telefone ---
        if "@lid" in remote_jid:
            telefone = remote_jid
            telefone_numero = remote_jid.split("@")[0]
        else:
            telefone = remote_jid.split("@")[0]
            telefone_numero = telefone

        nome_cliente = self._normalizar_nome_exibicao(body.data.pushName)
        id_mensagem = body.data.key.id
        contato_memoria_id = self._normalizar_contato_id(remote_jid)

        texto_recebido = self._extrair_texto(body)
        if not texto_recebido:
            return

        logger.info(f"[{instancia}][{telefone} / {contato_memoria_id} / {nome_cliente}]: {texto_recebido}")

        # --- Verifica blocklist/allowlist persistidos em banco, por instância ---
        if await self._is_blocked(instancia, telefone_numero):
            logger.info(f"Número %s está na blocklist da instância %s — ignorando.", telefone_numero, instancia)
            return

        if await self._has_allowlist_entries(instancia) and not await self._is_allowed(
            instancia, telefone_numero
        ):
            logger.info(
                "Número %s não está na allowlist da instância %s — ignorando.",
                telefone_numero,
                instancia,
            )
            return

        # --- Verifica estado da IA no banco (prioridade: por chat > global) ---
        ia_ativa = await self._ia_ativa_para(instancia, telefone_numero)
        if not ia_ativa:
            logger.info(f"IA desativada para [{instancia}][{telefone_numero}] — ignorando.")
            return

        # --- Salva mensagem recebida ---
        mensagem_nova = await self._salvar_mensagem(
            contato_memoria_id, nome_cliente, texto_recebido, "RECEBIDA", id_mensagem
        )
        if not mensagem_nova:
            logger.info(f"Mensagem duplicada ignorada: {id_mensagem}")
            return

        # --- Roteamento por instância: pessoal vs portfólio ---
        if instancia == settings.evolution_instance_two_name:
            await self._responder_instancia_pessoal(ev_client, telefone, contato_memoria_id, nome_cliente, texto_recebido)
        else:
            await self._responder_instancia_portfolio(ev_client, telefone, contato_memoria_id, nome_cliente, texto_recebido)

    # -----------------------------------------------------------------------
    # Fluxo da instância PORTFÓLIO (instância 1 — comportamento original)
    # -----------------------------------------------------------------------

    async def _responder_instancia_portfolio(
        self, ev_client: EvolutionClient, telefone: str, contato_memoria_id: str, nome: str, texto: str
    ):
        telefone_numero = telefone.split("@")[0] if "@" in telefone else telefone
        texto_lower = texto.lower()
        handled_document = await self._try_handle_document_request(ev_client, telefone, contato_memoria_id, nome, texto)
        if handled_document:
            return
        resposta_identidade = self._resposta_identidade_deterministica(texto)
        if resposta_identidade:
            try:
                await ev_client.send_text_message(telefone, resposta_identidade)
                await self._salvar_mensagem(contato_memoria_id, nome, resposta_identidade, "ENVIADA")
            except Exception as e:
                logger.error(f"Erro ao enviar resposta determinística para {telefone}: {e}")
            return

        _FORMATO_AUDIO = {"áudio", "audio", "voz", "voice"}
        _FORMATO_PLANILHA = {"planilha", "excel", "spreadsheet", "xlsx", "xls"}
        _VERBOS_ACAO = {
            "manda", "mandar", "envia", "enviar", "me manda", "me envia",
            "gera", "gerar", "cria", "criar", "faz", "fazer", "quero",
            "preciso", "pode", "consegue", "testa", "teste",
        }

        _quer_audio = any(k in texto_lower for k in _FORMATO_AUDIO)
        _quer_planilha = any(k in texto_lower for k in _FORMATO_PLANILHA) or (
            "tabela" in texto_lower and any(v in texto_lower for v in _VERBOS_ACAO)
        )

        _REMOVER_DA_QUERY = _FORMATO_AUDIO | _FORMATO_PLANILHA | {
            "me envia", "me manda", "me envie", "me mande",
            "manda", "envia", "gera", "cria", "faz",
            "em formato de", "em formato", "como", "no formato",
            "por favor", "pfv", "pf",
        }
        topico_query = texto_lower
        for rem in _REMOVER_DA_QUERY:
            topico_query = topico_query.replace(rem, " ")
        topico_query = " ".join(topico_query.split()).strip() or texto

        logger.info(f"RAG query topic: '{topico_query}' (audio={_quer_audio}, planilha={_quer_planilha})")
        await self.ensure_rag_ready()

        contexto_rag = await self.rag.retrieve_smart(topico_query)
        projeto_md = self.rag.load_project_if_mentioned(topico_query)
        contexto = self._combinar_contextos(
            self.rag.get_minimum_context(),
            projeto_md,
            contexto_rag,
        )

        historico_str = await self._obter_historico(contato_memoria_id, limite=8)
        contexto = self._aplicar_token_budget(contexto, historico_str, texto)

        if _quer_planilha:
            await self._responder_como_planilha(ev_client, telefone, nome, texto, contexto, historico_str)
            return
        if _quer_audio:
            await self._responder_como_audio_portfolio(ev_client, telefone, nome, texto, contexto, historico_str)
            return

        resposta_ia = await self._gerar_resposta_portfolio(nome, texto, contexto, historico_str)
        resposta_ia = self._remover_saudacao_repetida(resposta_ia, historico_str)
        try:
            await ev_client.send_text_message(telefone, resposta_ia)
            await self._salvar_mensagem(contato_memoria_id, nome, resposta_ia, "ENVIADA")
        except Exception as e:
            logger.error(f"Erro ao enviar resposta para {telefone}: {e}")

    # -----------------------------------------------------------------------
    # Fluxo da instância PESSOAL (instância 2 — Wesley informal)
    # -----------------------------------------------------------------------

    async def _responder_instancia_pessoal(
        self, ev_client: EvolutionClient, telefone: str, contato_memoria_id: str, nome: str, texto: str
    ):
        """Responde como o assistente pessoal do Wesley, usando o histórico da conversa."""
        texto_lower = texto.lower()
        handled_document = await self._try_handle_document_request(ev_client, telefone, contato_memoria_id, nome, texto)
        if handled_document:
            return
        resposta_identidade = self._resposta_identidade_deterministica(texto)
        if resposta_identidade:
            try:
                await ev_client.send_text_message(telefone, resposta_identidade)
                await self._salvar_mensagem(contato_memoria_id, nome, resposta_identidade, "ENVIADA")
            except Exception as e:
                logger.error(f"Erro ao enviar resposta determinística pessoal para {telefone}: {e}")
            return

        _FORMATO_AUDIO = {"áudio", "audio", "voz", "voice"}
        _quer_audio = any(k in texto_lower for k in _FORMATO_AUDIO)

        historico_str = await self._obter_historico(contato_memoria_id, limite=12)  # mais histórico para pegar o estilo
        
        # Recupera o contexto do portfólio para a instância pessoal também
        topico_query = " ".join(texto_lower.split()).strip() or texto
        await self.ensure_rag_ready()
        contexto_rag = await self.rag.retrieve_smart(topico_query)
        projeto_md = self.rag.load_project_if_mentioned(topico_query)
        contexto = self._combinar_contextos(
            self.rag.get_minimum_context(),
            projeto_md,
            contexto_rag,
        )
        contexto = self._aplicar_token_budget(contexto, historico_str, texto)

        if _quer_audio:
            resposta = await self._gerar_resposta_pessoal(nome, texto, historico_str, contexto, para_audio=True)
            resposta = self._remover_saudacao_repetida(resposta, historico_str)
            try:
                import asyncio
                def _gerar_tts():
                    tts = gTTS(text=resposta, lang="pt", slow=False)
                    audio_io = io.BytesIO()
                    tts.write_to_fp(audio_io)
                    audio_io.seek(0)
                    return base64.b64encode(audio_io.read()).decode("utf-8")
                
                b64 = await asyncio.wait_for(asyncio.to_thread(_gerar_tts), timeout=30.0)
                await ev_client.send_base64_audio(telefone, b64)
                await self._salvar_mensagem(contato_memoria_id, nome, resposta, "ENVIADA")
            except Exception as e:
                logger.error(f"Erro enviando áudio pessoal para {telefone}: {e}")
                # Fallback texto
                await ev_client.send_text_message(telefone, resposta)
                await self._salvar_mensagem(contato_memoria_id, nome, resposta, "ENVIADA")
        else:
            resposta = await self._gerar_resposta_pessoal(nome, texto, historico_str, contexto, para_audio=False)
            resposta = self._remover_saudacao_repetida(resposta, historico_str)
            try:
                await ev_client.send_text_message(telefone, resposta)
                await self._salvar_mensagem(contato_memoria_id, nome, resposta, "ENVIADA")
            except Exception as e:
                logger.error(f"Erro ao enviar resposta pessoal para {telefone}: {e}")

    async def _gerar_resposta_pessoal(
        self, nome: str, texto: str, historico: str, contexto: str, para_audio: bool = False
    ) -> str:
        """Gera resposta no estilo assistente pessoal cortês."""
        template = PROMPT_PESSOAL_AUDIO if para_audio else PROMPT_PESSOAL
        prompt = template.format(nome_cliente=nome, texto=texto, historico=historico, contexto=contexto)
        try:
            response = await self.llm_client.aio.models.generate_content(
                model=settings.gemini_model,
                contents=prompt,
            )
            texto_resp = response.text
            if para_audio:
                texto_resp = texto_resp.replace("**", "").replace("*", "").replace("#", "").replace("- ", "")
            return texto_resp
        except Exception as e:
            logger.error(f"Erro chamando IA (pessoal): {e}")
            return "ei, tô com uns problemas técnicos aqui, tenta de novo depois kkk"

    # -----------------------------------------------------------------------
    # Controle de IA via comandos /ia
    # -----------------------------------------------------------------------

    async def _processar_comando_ia(
        self,
        texto: str,
        instancia: str,
        owner_jid: str,
        ev_client: EvolutionClient,
    ) -> None:
        """Interpreta e executa um comando /ia enviado pelo owner."""
        partes = texto.strip().split()
        # partes[0] = "/ia", partes[1] = subcomando, partes[2] = número (opcional)

        if len(partes) < 2:
            await self._enviar_para_owner(ev_client, owner_jid, AJUDA_COMANDOS)
            return

        subcmd = partes[1].lower()          # on | off | lista | status | resetar
        numero_alvo = partes[2] if len(partes) >= 3 else None  # número opcional

        if subcmd in ("on", "off"):
            ativo = subcmd == "on"
            if numero_alvo:
                await self._set_config_ia(instancia, numero_alvo, ativo)
                emoji = "✅" if ativo else "🔴"
                msg = f"{emoji} IA {'ativada' if ativo else 'desativada'} para *{numero_alvo}*"
            else:
                # Config global (chat_jid = None)
                await self._set_config_ia(instancia, None, ativo)
                emoji = "✅" if ativo else "🔴"
                msg = f"{emoji} IA {'ativada' if ativo else 'desativada'} globalmente em *{instancia}*"
            logger.info(f"[COMANDO /ia] {msg}")
            await self._enviar_para_owner(ev_client, owner_jid, msg)

        elif subcmd == "resetar" and numero_alvo:
            await self._remover_config_chat(instancia, numero_alvo)
            msg = f"🔄 Override removido para *{numero_alvo}* — voltou ao padrão global de *{instancia}*"
            logger.info(f"[COMANDO /ia] {msg}")
            await self._enviar_para_owner(ev_client, owner_jid, msg)

        elif subcmd == "lista":
            resposta = await self._listar_conversas(instancia)
            await self._enviar_para_owner(ev_client, owner_jid, resposta)

        elif subcmd == "status":
            status_global = await self._ia_ativa_para(instancia, None)
            emoji = "✅" if status_global else "🔴"
            msg = f"{emoji} IA está *{'ATIVA' if status_global else 'DESATIVADA'}* globalmente em *{instancia}*"
            await self._enviar_para_owner(ev_client, owner_jid, msg)

        else:
            await self._enviar_para_owner(ev_client, owner_jid, AJUDA_COMANDOS)

    async def _enviar_para_owner(self, ev_client: EvolutionClient, owner_jid: str, msg: str) -> None:
        """Envia mensagem para o owner (número do dono do bot)."""
        if not owner_jid:
            logger.warning("owner_jid não configurado — não é possível enviar feedback de comando.")
            return
        # owner_jid pode já ter @s.whatsapp.net ou ser só o número
        destino = owner_jid if "@" in owner_jid else f"{owner_jid}@s.whatsapp.net"
        try:
            await ev_client.send_text_message(destino, msg)
        except Exception as e:
            logger.error(f"Erro ao enviar feedback de comando para owner {destino}: {e}")

    async def _listar_conversas(self, instancia: str) -> str:
        """Lista as 10 conversas mais recentes com status de IA."""
        async with async_session() as session:
            # Busca os 10 clientes com mensagem mais recente
            stmt = (
                select(Cliente, Mensagem)
                .join(Mensagem, Mensagem.id_cliente == Cliente.id)
                .order_by(Mensagem.data_hora.desc())
                .limit(20)
            )
            result = await session.execute(stmt)
            rows = result.all()

            # Deduplicar por cliente, mantendo o mais recente
            vistos: dict[str, tuple] = {}
            for cliente, mensagem in rows:
                if cliente.whatsapp_id not in vistos:
                    vistos[cliente.whatsapp_id] = (cliente, mensagem)
                if len(vistos) >= 10:
                    break

            if not vistos:
                return "📭 Nenhuma conversa registrada ainda."

            # Busca os configs de IA para esses JIDs
            jids = list(vistos.keys())
            stmt_cfg = select(BotConfig).where(
                BotConfig.instancia == instancia,
                or_(BotConfig.chat_jid.in_(jids), BotConfig.chat_jid.is_(None)),
            )
            result_cfg = await session.execute(stmt_cfg)
            configs = result_cfg.scalars().all()

            cfg_por_chat: dict[str | None, bool] = {}
            for c in configs:
                cfg_por_chat[c.chat_jid] = c.ia_ativa

            global_ativo = cfg_por_chat.get(None, True)  # padrão: ativo

            linhas = [f"📋 *Últimas conversas em {instancia}:*\n"]
            for jid, (cliente, mensagem) in vistos.items():
                # Prioridade: config individual > global
                telefone_numero = jid.split("@")[0] if "@" in jid else jid
                ativo = cfg_por_chat.get(telefone_numero, global_ativo)
                emoji = "✅" if ativo else "🔴"
                nome = cliente.nome or "Desconhecido"
                ultima = mensagem.data_hora.strftime("%d/%m %H:%M") if mensagem.data_hora else "—"
                linhas.append(f"{emoji} *{nome}* | `{telefone_numero}` | {ultima}")

            return "\n".join(linhas)

    # -----------------------------------------------------------------------
    # Persistência BotConfig
    # -----------------------------------------------------------------------

    async def _set_config_ia(self, instancia: str, chat_jid: Optional[str], ativo: bool) -> None:
        """Upsert: cria ou atualiza a config de IA para a instância/chat."""
        async with async_session() as session:
            stmt = select(BotConfig).where(
                BotConfig.instancia == instancia,
                BotConfig.chat_jid == chat_jid if chat_jid else BotConfig.chat_jid.is_(None),
            )
            result = await session.execute(stmt)
            config = result.scalar_one_or_none()

            if config:
                config.ia_ativa = ativo
                config.updated_at = datetime.utcnow()
            else:
                config = BotConfig(
                    id=str(uuid.uuid4()),
                    instancia=instancia,
                    chat_jid=chat_jid,
                    ia_ativa=ativo,
                    updated_at=datetime.utcnow(),
                )
                session.add(config)

            await session.commit()

    async def _remover_config_chat(self, instancia: str, chat_jid: str) -> None:
        """Remove o override individual de um chat específico."""
        async with async_session() as session:
            stmt = select(BotConfig).where(
                BotConfig.instancia == instancia,
                BotConfig.chat_jid == chat_jid,
            )
            result = await session.execute(stmt)
            config = result.scalar_one_or_none()
            if config:
                await session.delete(config)
                await session.commit()

    async def _ia_ativa_para(self, instancia: str, chat_jid: Optional[str]) -> bool:
        """Consulta o BD para saber se a IA está ativa para este chat.
        
        Prioridade: config individual > config global > padrão True.
        """
        async with async_session() as session:
            # Busca config individual do chat (se chat_jid fornecido)
            if chat_jid:
                stmt_chat = select(BotConfig).where(
                    BotConfig.instancia == instancia,
                    BotConfig.chat_jid == chat_jid,
                )
                result = await session.execute(stmt_chat)
                cfg_chat = result.scalar_one_or_none()
                if cfg_chat is not None:
                    return cfg_chat.ia_ativa

            # Fallback: config global da instância
            stmt_global = select(BotConfig).where(
                BotConfig.instancia == instancia,
                BotConfig.chat_jid.is_(None),
            )
            result = await session.execute(stmt_global)
            cfg_global = result.scalar_one_or_none()
            if cfg_global is not None:
                return cfg_global.ia_ativa

            return True  # padrão: IA ativa

    # -----------------------------------------------------------------------
    # Allowlist / Blocklist persistidos em banco (AllowBlockEntry)
    # -----------------------------------------------------------------------

    async def _is_blocked(self, instancia: str, numero: str) -> bool:
        """Retorna True se o número estiver na blocklist dessa instância."""
        async with async_session() as session:
            stmt = select(AllowBlockEntry).where(
                AllowBlockEntry.tipo == "block",
                AllowBlockEntry.numero == numero,
                AllowBlockEntry.instancia == instancia,
            )
            result = await session.execute(stmt)
            return result.scalar_one_or_none() is not None

    async def _has_allowlist_entries(self, instancia: str) -> bool:
        """Retorna True se existir qualquer allowlist configurada para a instância."""
        async with async_session() as session:
            stmt = select(AllowBlockEntry).where(
                AllowBlockEntry.tipo == "allow",
                AllowBlockEntry.instancia == instancia,
            )
            result = await session.execute(stmt)
            return result.first() is not None

    async def _is_allowed(self, instancia: str, numero: str) -> bool:
        """Retorna True se o número estiver na allowlist dessa instância."""
        async with async_session() as session:
            stmt = select(AllowBlockEntry).where(
                AllowBlockEntry.tipo == "allow",
                AllowBlockEntry.numero == numero,
                AllowBlockEntry.instancia == instancia,
            )
            result = await session.execute(stmt)
            return result.scalar_one_or_none() is not None

    # -----------------------------------------------------------------------
    # Geração de respostas — Portfólio
    # -----------------------------------------------------------------------

    async def _gerar_resposta_portfolio(
        self, nome: str, texto: str, contexto: str, historico: str = ""
    ) -> str:
        historico_prompt = (
            f"HISTÓRICO RECENTE DA CONVERSA:\n{historico}\n"
            "Use o histórico para manter a fluidez da conversa, mas FOQUE a sua resposta "
            "principalmente na última mensagem abaixo.\n\n"
        ) if historico else ""
        prompt = PROMPT_PORTFOLIO.format(
            nome_cliente=nome,
            contexto=contexto,
            historico_prompt=historico_prompt,
            texto=texto,
        )
        try:
            response = await self.llm_client.aio.models.generate_content(
                model=settings.gemini_model, contents=prompt
            )
            return response.text
        except Exception as e:
            logger.error(f"Erro chamando IA: {e}")
            return "Ops, dei uma travadinha processando seu portfólio. Manda de novo?"

    async def _gerar_resposta_portfolio_audio(
        self, nome: str, texto: str, contexto: str, historico: str = ""
    ) -> str:
        historico_prompt = (
            f"HISTÓRICO RECENTE DA CONVERSA:\n{historico}\n"
            "Use o histórico para manter a fluidez da conversa, mas FOQUE a sua resposta "
            "principalmente na última mensagem abaixo.\n\n"
        ) if historico else ""
        prompt = PROMPT_PORTFOLIO_AUDIO.format(
            nome_cliente=nome,
            contexto=contexto,
            historico_prompt=historico_prompt,
            texto=texto,
        )
        try:
            response = await self.llm_client.aio.models.generate_content(
                model=settings.gemini_model, contents=prompt
            )
            clean = response.text.replace("**", "").replace("*", "").replace("#", "").replace("- ", "")
            return clean
        except Exception as e:
            logger.error(f"Erro chamando IA: {e}")
            return "Putz, deu uma travadinha aqui na hora de carregar meu cérebro. Tenta me pedir de novo?"

    async def _responder_como_audio_portfolio(
        self,
        ev_client: EvolutionClient,
        telefone: str,
        nome: str,
        texto: str,
        contexto: str,
        historico: str = "",
    ):
        resposta_texto = await self._gerar_resposta_portfolio_audio(nome, texto, contexto, historico)
        resposta_texto = self._remover_saudacao_repetida(resposta_texto, historico)
        logger.info(f"Gerando áudio TTS para {telefone}: {resposta_texto[:60]}...")
        try:
            import asyncio
            def _gerar_tts_portfolio():
                tts = gTTS(text=resposta_texto, lang="pt", slow=False)
                audio_io = io.BytesIO()
                tts.write_to_fp(audio_io)
                audio_io.seek(0)
                return base64.b64encode(audio_io.read()).decode("utf-8")

            b64 = await asyncio.wait_for(asyncio.to_thread(_gerar_tts_portfolio), timeout=30.0)
            await ev_client.send_base64_audio(telefone, b64)
            await self._salvar_mensagem(self._normalizar_contato_id(telefone), nome, resposta_texto, "ENVIADA")
        except Exception as e:
            logger.error(f"Erro enviando áudio para {telefone}: {e}")
            await ev_client.send_text_message(telefone, resposta_texto)
            await self._salvar_mensagem(self._normalizar_contato_id(telefone), nome, resposta_texto, "ENVIADA")

    async def _responder_como_planilha(
        self,
        ev_client: EvolutionClient,
        telefone: str,
        nome: str,
        texto: str,
        contexto: str,
        historico: str = "",
    ):
        prompt_planilha = f"""
Você é o Assistente do Wesley. O usuário pediu: "{texto}"

Com base neste contexto do portfólio do Wesley:
{contexto}

Extraia as informações pedidas e retorne SOMENTE as linhas correspondentes ao pedido do usuário (ignore informações secundárias de RAG), sem texto extra.
Formato obrigatório: cada linha separada por \\n, colunas separadas por | (pipe).
A primeira linha é o cabeçalho. Exemplo:
Tecnologia | Nível | Categoria
Python | Avançado | Backend
"""
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side
            response = await self.llm_client.aio.models.generate_content(
                model=settings.gemini_model, contents=prompt_planilha
            )
            linhas_raw = response.text.strip().split("\n")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Portfolio Wesley"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            thin = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            row_idx = 1
            for linha in linhas_raw:
                if "|" in linha:
                    colunas = [c.strip() for c in linha.split("|")]
                    ws.append(colunas)
                    for col_idx in range(1, len(colunas) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin
                        if row_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                    row_idx += 1

            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

            excel_io = io.BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)
            b64 = base64.b64encode(excel_io.read()).decode("utf-8")
            caption = f"Planilha gerada para {nome} com base no portfólio do Wesley!"
            await ev_client.send_base64_document(telefone, b64, "Wesley_Portfolio.xlsx", caption)
            await self._salvar_mensagem(self._normalizar_contato_id(telefone), nome, "[Planilha Excel gerada e enviada]", "ENVIADA")
        except Exception as e:
            logger.error(f"Erro enviando planilha para {telefone}: {e}")
            resposta = await self._gerar_resposta_portfolio(nome, texto, contexto, historico)
            resposta = self._remover_saudacao_repetida(resposta, historico)
            await ev_client.send_text_message(telefone, resposta)
            await self._salvar_mensagem(self._normalizar_contato_id(telefone), nome, resposta, "ENVIADA")

    # -----------------------------------------------------------------------
    # Utilidades
    # -----------------------------------------------------------------------

    def _extrair_texto(self, body: WebhookBody) -> Optional[str]:
        msg_obj = body.data.message
        if not msg_obj:
            return None
        if msg_obj.conversation:
            return msg_obj.conversation
        if msg_obj.extendedTextMessage and "text" in msg_obj.extendedTextMessage:
            return msg_obj.extendedTextMessage["text"]
        return None

    def _aplicar_token_budget(
        self, contexto: str, historico: str, texto: str, max_tokens: int = 30000
    ) -> str:
        max_chars = max_tokens * 4
        base_chars = len(historico) + len(texto) + 1000
        if base_chars + len(contexto) <= max_chars:
            return contexto
        chars_disponiveis = max_chars - base_chars
        if chars_disponiveis <= 0:
            return ""
        logger.warning(f"Token Budget: cortando RAG de {len(contexto)} para {chars_disponiveis} chars.")
        return contexto[:chars_disponiveis]

    def _normalizar_contato_id(self, whatsapp_id: Optional[str]) -> str:
        if not whatsapp_id:
            return ""
        base = whatsapp_id.split("@")[0].strip()
        digits = "".join(ch for ch in base if ch.isdigit())
        return digits or base

    def _variantes_contato_id(self, whatsapp_id: Optional[str]) -> list[str]:
        variantes: list[str] = []
        bruto = (whatsapp_id or "").strip()
        normalizado = self._normalizar_contato_id(bruto)
        for valor in (bruto, normalizado):
            if valor and valor not in variantes:
                variantes.append(valor)
        if normalizado:
            for sufixo in ("@s.whatsapp.net", "@lid"):
                variante = f"{normalizado}{sufixo}"
                if variante not in variantes:
                    variantes.append(variante)
        return variantes

    def _normalizar_nome_exibicao(self, nome: Optional[str]) -> str:
        clean = " ".join((nome or "").split()).strip()
        if not clean:
            return "contato"
        primeiro_nome = clean.split()[0]
        return primeiro_nome[:40]

    def _normalizar_texto_intencao(self, texto: str) -> str:
        texto = unicodedata.normalize("NFKD", texto.lower())
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        return " ".join(texto.split())

    def _detectar_idioma_documento(self, texto: str) -> str:
        texto_norm = self._normalizar_texto_intencao(texto)
        if any(k in texto_norm for k in ("english", "in english", "ingles", "em ingles", "resume")):
            return "en"
        return "pt"

    def _is_certificate_listing_request(self, texto: str) -> bool:
        texto_norm = self._normalizar_texto_intencao(texto)
        return (
            any(term in texto_norm for term in ("certificado", "certificados", "certificate", "certificates"))
            and any(
                termo in texto_norm
                for termo in ("quais", "listar", "lista", "tenho", "possuo", "mostrar", "mostra")
            )
        )

    def _is_send_request(self, texto: str) -> bool:
        texto_norm = self._normalizar_texto_intencao(texto)
        return any(
            termo in texto_norm
            for termo in (
                "manda", "manda ai", "mandar", "envia", "enviar", "me manda",
                "me envia", "pode mandar", "pode enviar", "send", "attach", "anexa",
            )
        )

    def _build_document_caption(self, entry: DocumentEntry, language: str) -> str:
        if entry.category == "resume_en":
            return "Segue o resume em inglês do Wesley."
        if entry.category == "resume_pt":
            return "Segue o currículo do Wesley."
        return (
            f"Segue o certificado: {entry.title_en}."
            if language == "en"
            else f"Segue o certificado: {entry.title_pt}."
        )

    async def _try_handle_document_request(
        self,
        ev_client: EvolutionClient,
        telefone: str,
        contato_memoria_id: str,
        nome: str,
        texto: str,
    ) -> bool:
        language = self._detectar_idioma_documento(texto)

        if self._is_certificate_listing_request(texto):
            resposta = self.document_catalog.build_certificate_list_message(language)
            await ev_client.send_text_message(telefone, resposta)
            await self._salvar_mensagem(contato_memoria_id, nome, resposta, "ENVIADA")
            return True

        if not self._is_send_request(texto):
            return False

        entry = self.document_catalog.find_best_document(texto, language)
        if not entry:
            return False

        b64 = self.document_catalog.load_base64(entry)
        caption = self._build_document_caption(entry, language)
        await ev_client.send_base64_document(telefone, b64, entry.filename, caption)
        await self._salvar_mensagem(contato_memoria_id, nome, f"[Documento enviado] {entry.filename}", "ENVIADA")
        return True

    def _resposta_identidade_deterministica(self, texto: str) -> Optional[str]:
        texto_norm = self._normalizar_texto_intencao(texto)

        pergunta_nome_assistente = (
            ("qual seu nome" in texto_norm or "seu nome" in texto_norm or "quem e voce" in texto_norm)
            and "wesley" not in texto_norm
        )
        if pergunta_nome_assistente:
            return f"Eu sou {ASSISTANT_DISPLAY_NAME} aqui no WhatsApp."

        pergunta_nome_completo = (
            "nome completo" in texto_norm
            or "nome inteiro" in texto_norm
            or "sobrenome" in texto_norm
            or "nome de verdade" in texto_norm
            or "nome completo de verdade" in texto_norm
        )
        if pergunta_nome_completo:
            return f"O nome completo dele é {WESLEY_PUBLIC_FULL_NAME}."

        pergunta_nome_wesley = (
            "nome do wesley" in texto_norm
            or "quem e o wesley" in texto_norm
            or ("nome dele" in texto_norm and "wesley" in texto_norm)
        )
        if pergunta_nome_wesley:
            return f"O nome dele é {WESLEY_PUBLIC_NAME}."

        return None

    def _combinar_contextos(self, *partes: Optional[str]) -> str:
        partes_unicas: list[str] = []
        for parte in partes:
            trecho = (parte or "").strip()
            if trecho and trecho not in partes_unicas:
                partes_unicas.append(trecho)
        return "\n---\n".join(partes_unicas)

    def _historico_tem_saudacao_anterior(self, historico: str) -> bool:
        historico_lower = historico.lower()
        return any(
            marcador in historico_lower
            for marcador in (
                "assistente: olá",
                "assistente: ola",
                "assistente: oi",
                "assistente: opa",
                "assistente: tudo bem",
            )
        )

    def _remover_saudacao_repetida(self, resposta: str, historico: str) -> str:
        if not resposta or not self._historico_tem_saudacao_anterior(historico):
            return resposta
        texto = resposta.strip()
        padroes = [
            r"^(oi|olá|ola|opa|e aí|eaí|ei)\b[^.!?\n]*[.!?]\s*",
            r"^tudo bem\??\s*",
            r"^(oi|olá|ola|opa)\s+[^\n,!?]{1,40},?\s*",
        ]
        for padrao in padroes:
            novo = re.sub(padrao, "", texto, flags=re.IGNORECASE)
            if novo != texto:
                texto = novo.strip()
        return texto or resposta

    async def _salvar_mensagem(
        self,
        whatsapp_id: str,
        nome: str,
        texto: str,
        direcao: str,
        msg_id: Optional[str] = None,
    ) -> bool:
        if not texto:
            return False
        async with async_session() as session:
            contato_id = self._normalizar_contato_id(whatsapp_id)
            variantes = self._variantes_contato_id(whatsapp_id)
            stmt = select(Cliente).where(Cliente.whatsapp_id.in_(variantes))
            result = await session.execute(stmt)
            cliente = result.scalar_one_or_none()

            if not cliente:
                cliente = Cliente(id=str(uuid.uuid4()), whatsapp_id=contato_id or whatsapp_id, nome=nome)
                session.add(cliente)
                await session.flush()
            elif not cliente.nome and nome:
                cliente.nome = nome

            if msg_id:
                stmt_msg = select(Mensagem).where(Mensagem.mensagem_id_whatsapp == msg_id)
                result_msg = await session.execute(stmt_msg)
                if result_msg.scalar_one_or_none():
                    return False

            nova_msg = Mensagem(
                id=str(uuid.uuid4()),
                id_cliente=cliente.id,
                texto=texto,
                mensagem_id_whatsapp=msg_id or str(uuid.uuid4()),
                direcao=direcao,
                data_hora=datetime.utcnow(),
            )
            session.add(nova_msg)
            await session.commit()
            return True

    async def _obter_historico(self, whatsapp_id: str, limite: int = 5) -> str:
        async with async_session() as session:
            stmt = select(Cliente).where(Cliente.whatsapp_id.in_(self._variantes_contato_id(whatsapp_id)))
            result = await session.execute(stmt)
            cliente = result.scalar_one_or_none()
            if not cliente:
                return ""

            stmt_msg = (
                select(Mensagem)
                .where(Mensagem.id_cliente == cliente.id)
                .order_by(Mensagem.data_hora.desc())
                .limit(limite)
            )
            result_msg = await session.execute(stmt_msg)
            mensagens = result_msg.scalars().all()

            if not mensagens:
                return ""

            historico = []
            for m in reversed(mensagens):
                remetente = "Usuário" if m.direcao == "RECEBIDA" else "Assistente"
                historico.append(f"{remetente}: {m.texto}")
            return "\n".join(historico)
